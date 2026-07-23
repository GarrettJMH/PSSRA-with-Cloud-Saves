
# Importing Streamlit.
import streamlit as st

# Importing pandas.
import pandas as pd

# Importing date and timedelta for deadline calculations and calendar schedule display.
from datetime import date, timedelta

# Importing math to help with urgency scheduling.
import math

import re

# -----------------------------
# GENERAL HELPERS
# -----------------------------

def split_items(value):
    """
    Converts a comma-separated string into a list.
    Also safely handles lists, None values, and blank/NaN spreadsheet cells.
    """

    # If the value is already a list, return it directly.
    if isinstance(value, list):
        return value

    # If the value is None/empty, return an empty list.
    if value is None:
        return []

    # Handles blank cells from CSV/Excel imports, which may appear as NaN.
    if pd.isna(value):
        return []

    # Convert the value to a string, split by commas, remove spaces,
    # and ignore any blank entries.
    return [
        item.strip()
        for item in str(value).split(",")
        if item.strip() != ""
    ]


def is_blank_value(value):
    """
    Returns True if a value should be treated as blank.
    Handles None, empty strings, and NaN values from pandas.
    """

    if value is None:
        return True

    try:
        if pd.isna(value):
            return True
    except (TypeError, ValueError):
        pass

    return str(value).strip() == ""


def clean_records(records):
    """
    Removes empty rows that may appear from editable tables.
    Treats None, blank strings, and NaN values as empty.
    """

    cleaned = []

    for record in records:
        has_content = False

        for value in record.values():
            if not is_blank_value(value):
                has_content = True
                break

        if has_content:
            cleaned.append(record)

    return cleaned


def split_unavailability_items(value):
    """
    Splits the Unavailability field into entries.

    Supports:
    - Week numbers: 3, 4, 8
    - Date ranges: 2026-08-10 to 2026-08-14
    - Single dates: 2026-08-21
    - Mixed input: 3, 4; 2026-08-10 to 2026-08-14
    """

    if is_blank_value(value):
        return []

    text = str(value).replace(";", ",")

    return [
        item.strip()
        for item in text.split(",")
        if item.strip() != ""
    ]


def parse_unavailability_date_item(item):
    """
    Parses one unavailability item as either a single date or a date range.

    Examples:
    2026-08-10
    2026-08-10 to 2026-08-14
    """

    item = str(item).strip()

    # Split date ranges written as "date to date".
    parts = re.split(r"\s+to\s+", item, flags=re.IGNORECASE)

    try:
        if len(parts) == 1:
            start_date = pd.to_datetime(parts[0]).date()
            end_date = start_date

        elif len(parts) == 2:
            start_date = pd.to_datetime(parts[0]).date()
            end_date = pd.to_datetime(parts[1]).date()

        else:
            return None

    except Exception:
        return None

    # If the user accidentally enters the dates backward, fix the order.
    if end_date < start_date:
        start_date, end_date = end_date, start_date

    return start_date, end_date

def parse_week_number(item):
    """
    Parses a week number from text or spreadsheet-style numeric values.

    Accepts:
    - 3
    - "3"
    - 3.0
    - "3.0"

    Rejects:
    - 3.5
    - dates
    - blank values
    """

    try:
        number = float(str(item).strip())
    except (TypeError, ValueError):
        return None

    if number.is_integer() and number >= 1:
        return int(number)

    return None

def date_to_model_week(input_date, planning_start_date):
    """
    Converts a real date into the model week number.
    Week 1 begins on planning_start_date.
    """

    if input_date < planning_start_date:
        return None

    days_from_start = (input_date - planning_start_date).days

    return (days_from_start // 7) + 1


def value_is_true(value):
    """
    Converts checkbox/text values into a reliable True/False value.
    This helps because imported Excel/CSV values may be True, FALSE, yes, 1, etc.
    """

    # Empty values should not count as True.
    if value is None:
        return False

    # Blank spreadsheet cells may appear as NaN.
    if pd.isna(value):
        return False

    # Accept several common ways of writing a true value.
    return str(value).strip().lower() in ["true", "yes", "y", "1"]


# -----------------------------
# FILE IMPORT HELPERS
# -----------------------------

def read_uploaded_table(uploaded_file):
    """
    Reads either a CSV or Excel file into a Pandas DataFrame.
    """

    # Convert the filename to lowercase so extension checking is consistent.
    file_name = uploaded_file.name.lower()

    # Read CSV files.
    if file_name.endswith(".csv"):
        return pd.read_csv(uploaded_file)

    # Read Excel files.
    elif file_name.endswith(".xlsx"):
        return pd.read_excel(uploaded_file)

    # If the file is not supported, show an error and return None.
    else:
        st.error("Unsupported file type. Please upload a CSV or Excel file.")
        return None


# Loads project data from an uploaded CSV or Excel file.
def load_projects_from_file(uploaded_file):
    """
    Loads project data from an uploaded CSV or Excel file.
    Required columns: Project name, Priority, Roles.
    Optional columns: Mandatory, Depends on, Conflicts with, Uploaded file.
    """

    # Read the uploaded file into a DataFrame.
    projects_df = read_uploaded_table(uploaded_file)

    # Stop if the file could not be read.
    if projects_df is None:
        return []

    # These columns must exist for the optimizer to understand projects.
    required_columns = ["Project name", "Priority", "Roles"]

    # Check that all required columns exist.
    for column in required_columns:
        if column not in projects_df.columns:
            st.error(f"Missing required column in project file: {column}")
            return []

    # If the file does not include a Mandatory column, assume projects are not mandatory by default.
    if "Mandatory" not in projects_df.columns:
        projects_df["Mandatory"] = False

    # If the file does not include deadline information, use a blank deadline so older project files still work.
    if "Deadline" not in projects_df.columns:
        projects_df["Deadline"] = ""

    if "Project context" not in projects_df.columns:
        projects_df["Project context"] = ""

    # If the file does not include duration information, assume each project takes one week by default.
    if "Estimated duration (weeks)" not in projects_df.columns:
        projects_df["Estimated duration (weeks)"] = 1

    # If the file does not include default role-hour information, assume each required role uses one hour per week by default.
    if "Role hours/week" not in projects_df.columns:
        projects_df["Role hours/week"] = 1

    # If the file does not include dependency information, use blank dependency values by default.
    if "Depends on" not in projects_df.columns:
        projects_df["Depends on"] = ""

    # If the file does not include conflict information, use blank conflict values by default.
    if "Conflicts with" not in projects_df.columns:
        projects_df["Conflicts with"] = ""

    # Track the source file for imported project rows.
    if "Uploaded file" not in projects_df.columns:
        projects_df["Uploaded file"] = uploaded_file.name

    # Convert the DataFrame into a list of dictionaries and remove blank rows.
    return clean_records(projects_df.to_dict("records"))


# Loads worker data from an uploaded CSV or Excel file.
def load_workers_from_file(uploaded_file):
    """
    Loads worker data from an uploaded CSV or Excel file.
    Required columns: Worker name, Profile, Skills, Weekly hours.
    Optional column: Uploaded profile.
    """

    # Read the uploaded file into a DataFrame.
    workers_df = read_uploaded_table(uploaded_file)

    # Stop if the file could not be read.
    if workers_df is None:
        return []

    # These columns must exist for the optimizer and Q generator to understand workers.
    required_columns = ["Worker name", "Profile", "Skills", "Weekly hours"]

    # Check that all required columns exist.
    for column in required_columns:
        if column not in workers_df.columns:
            st.error(f"Missing required column in worker file: {column}")
            return []

    # Track the source file for imported worker rows.
    if "Uploaded profile" not in workers_df.columns:
        workers_df["Uploaded profile"] = uploaded_file.name

    # Check the file for preferred roles.
    if "Preferred roles" not in workers_df.columns:
        workers_df["Preferred roles"] = ""

    # Check for unavailability. If missing, the worker is assumed to be available every week.
    if "Unavailability" not in workers_df.columns:
        workers_df["Unavailability"] = ""

    #Checking the uploaded file has the required weekly hours column.
    if "Weekly hours" not in workers_df.columns:
        st.error("Missing required column in worker file: Weekly hours")
        return []

    # Convert the DataFrame into a list of dictionaries and remove blank rows.
    return clean_records(workers_df.to_dict("records"))


# -----------------------------
# DATE AND SCHEDULING HELPERS
# -----------------------------

# Calculates real deadline dates into numeric values that can be used by the optimizer.
def calculate_deadline(deadline_value, planning_start_date=None):
    """
    Converts a real project deadline into:
    - days until due, for display/debugging
    - weeks until due, for scheduling
    - deadline urgency score, for effective priority calculations
    """

    if deadline_value is None or str(deadline_value).strip() == "":
        return None, None, 0

    try:
        deadline = pd.to_datetime(deadline_value).date()
    except Exception:
        return None, None, 0

    if planning_start_date is None:
        planning_start_date = date.today()

    days_till_due = (deadline - planning_start_date).days
    weeks_till_due = math.ceil(days_till_due / 7)

    if weeks_till_due <= 1:
        urgency = 4
    elif weeks_till_due <= 2:
        urgency = 3
    elif weeks_till_due <= 3:
        urgency = 2
    elif weeks_till_due <= 5:
        urgency = 1
    else:
        urgency = 0

    return days_till_due, weeks_till_due, urgency


def week_to_date_range(week_number, planning_start_date):
    """
    Converts a model week number into a calendar date range.
    Week 1 starts on the planning_start_date.
    """

    start_date = planning_start_date + timedelta(days=(week_number - 1) * 7)
    end_date = start_date + timedelta(days=6)

    return f"{start_date} to {end_date}"


# -----------------------------
# PROJECT DATA BUILDERS
# -----------------------------

def parse_specific_role_hours(value):
    """
    Converts a text value like:
    Developer: 8, Tester: 3, Writer: 4

    into a dictionary like:
    {
        "Developer": 8.0,
        "Tester": 3.0,
        "Writer": 4.0
    }
    """

    if isinstance(value, dict):
        return {
            str(role).strip(): float(hours)
            for role, hours in value.items()
            if str(role).strip() != ""
        }

    role_hours = {}

    if value is None:
        return role_hours

    if pd.isna(value):
        return role_hours

    entries = str(value).split(",")

    for entry in entries:
        if ":" in entry:
            role, hours = entry.split(":", 1)

            role = role.strip()

            try:
                hours = float(hours.strip())
                role_hours[role] = hours
            except ValueError:
                pass

    return role_hours


def build_mandatory_projects(project_rows):
    """
    Builds a list of projects that must be selected by the optimizer.
    """

    mandatory_projects = []

    # Loop through each project row.
    for row in project_rows:
        # Get and clean the project name.
        project_name = str(row.get("Project name", "")).strip()

        # Skip blank project names.
        if project_name == "":
            continue

        # If the Mandatory column is true, add the project to the mandatory list.
        if value_is_true(row.get("Mandatory", False)):
            mandatory_projects.append(project_name)

    return mandatory_projects


# Builds project dependency relationships.
def build_project_dependencies(project_rows):
    """
    Builds dependency pairs in the format:
    (dependent_project, required_project)

    Example:
    Firewall upgrade depends on Software maintenance update

    becomes:
    ("Firewall upgrade", "Software maintenance update")
    """

    project_dependencies = []

    # Create a set of valid project names.
    project_names = {
        str(row.get("Project name", "")).strip()
        for row in project_rows
        if str(row.get("Project name", "")).strip() != ""
    }

    # Loop through each project row.
    for row in project_rows:
        # This project is the dependent project.
        dependent_project = str(row.get("Project name", "")).strip()

        # Skip blank project names.
        if dependent_project == "":
            continue

        # Read the projects this project depends on.
        required_projects = split_items(row.get("Depends on", ""))

        # Create one dependency pair for each required project.
        for required_project in required_projects:
            required_project = str(required_project).strip()

            if (
                required_project != ""
                and required_project != dependent_project
                and required_project in project_names
            ):
                project_dependencies.append((dependent_project, required_project))

    return project_dependencies


# Builds project conflict relationships.
def build_project_conflicts(project_rows):
    """
    Builds conflict pairs in the format:
    (project_a, project_b)

    Example:
    Firewall upgrade conflicts with Fibre rerouting

    becomes:
    ("Firewall upgrade", "Fibre rerouting")
    """

    project_conflicts = []

    # Create a set of valid project names.
    project_names = {
        str(row.get("Project name", "")).strip()
        for row in project_rows
        if str(row.get("Project name", "")).strip() != ""
    }

    # Loop through each project row.
    for row in project_rows:
        # This is the first project in the conflict pair.
        project_a = str(row.get("Project name", "")).strip()

        # Skip blank project names.
        if project_a == "":
            continue

        # Read the projects that conflict with this project.
        conflicting_projects = split_items(row.get("Conflicts with", ""))

        # Create one conflict pair for each conflicting project.
        for project_b in conflicting_projects:
            project_b = str(project_b).strip()

            if (
                project_b != ""
                and project_b != project_a
                and project_b in project_names
            ):
                # Sorting prevents duplicate conflict pairs in opposite order.
                pair = tuple(sorted([project_a, project_b]))

                if pair not in project_conflicts:
                    project_conflicts.append(pair)

    return project_conflicts



def parse_workers_per_role(value):
    """
    Converts text like:
    Developer: 2, Tester: 1

    into a dictionary like:
    {
        "Developer": 2,
        "Tester": 1
    }
    """

    if isinstance(value, dict):
        cleaned = {}

        for role, count in value.items():
            role = str(role).strip()

            if role == "":
                continue

            try:
                cleaned[role] = max(1, int(float(count)))
            except (TypeError, ValueError):
                cleaned[role] = 1

        return cleaned

    role_requirements = {}

    if value is None:
        return role_requirements

    if pd.isna(value):
        return role_requirements

    entries = str(value).split(",")

    for entry in entries:
        if ":" in entry:
            role, count = entry.split(":", 1)
            role = role.strip()

            if role == "":
                continue

            try:
                role_requirements[role] = max(1, int(float(count.strip())))
            except ValueError:
                pass

    return role_requirements

def build_projects(project_rows, planning_start_date=None):
    # Dictionary that will store all projects for the optimization model.
    projects = {}

    # Loop through each project row entered in the Streamlit app.
    for row in project_rows:
        # Get and clean the project name from the row.
        raw_project_name = row.get("Project name", "")

        # Skip blank project names, including NaN from editable tables.
        if is_blank_value(raw_project_name):
            continue

        project_name = str(raw_project_name).strip()

        # Get the selected roles for the project.
        required_roles = split_items(row.get("Roles", ""))

        # Skip rows with no roles.
        if len(required_roles) == 0:
            continue

        # Get the base project priority safely.
        try:
            base_priority = int(float(row.get("Priority", 1)))
        except (TypeError, ValueError):
            st.warning(f"Skipping project '{project_name}' because its priority is invalid.")
            continue

        # Get and process the project deadline.
        deadline = row.get("Deadline", "")
        days_till_due, weeks_till_due, urgency = calculate_deadline(deadline, planning_start_date)

        # Get the estimated project duration safely.
        try:
            estimated_duration = int(float(row.get("Estimated duration (weeks)", 1)))
        except (TypeError, ValueError):
            estimated_duration = 1

        # Combine base priority and deadline urgency to create a deadline-aware priority score.
        effective_priority = base_priority + urgency

        # Build all valid start weeks for the scheduling model.
        # A project is deadline-feasible only if it can start early enough to finish by its deadline.
        if weeks_till_due is None:
            deadline_feasible = True
            deadline_note = ""
            valid_start_weeks = [1]

        else:
            latest_start_week = weeks_till_due - estimated_duration + 1

            if latest_start_week >= 1:
                deadline_feasible = True
                deadline_note = ""
                valid_start_weeks = list(range(1, latest_start_week + 1))

            else:
                deadline_feasible = False
                deadline_note = (
                    f"The project requires {estimated_duration} week(s), but only "
                    f"{weeks_till_due} week(s) are available before the deadline."
                )

                # Keep a placeholder start week so the optimizer can still build variables safely.
                # The optimizer will force this project to be unselected.
                valid_start_weeks = [1]

        # Get the default weekly hours required for each role on this project safely.
        try:
            role_hours = float(row.get("Role hours/week", 1))
        except (TypeError, ValueError):
            role_hours = 1

        # Get optional specific weekly hours for each role in the project.
        specific_role_hours = parse_specific_role_hours(
            row.get("Role hours per week", "")
        )

        # Get the number of workers needed for each project role.
        role_requirements = parse_workers_per_role(
            row.get("Workers per role", "")
        )

        # If a selected role does not have an explicit worker count, default to one worker.
        for role in required_roles:
            if role not in role_requirements:
                role_requirements[role] = 1

        # Store each project using the format needed for the model.
        projects[project_name] = {
            "project_context": row.get("Project context", ""),
            "priority": base_priority,
            "deadline": deadline,
            "days_until_due": days_till_due,
            "weeks_until_due": weeks_till_due,
            "deadline_urgency": urgency,
            "deadline_feasible": deadline_feasible,
            "deadline_note": deadline_note,
            "effective_priority": effective_priority,
            "estimated_duration_weeks": estimated_duration,
            "valid_start_weeks": valid_start_weeks,
            "role_hours_per_week": role_hours,
            "specific_role_hours": specific_role_hours,
            "role_requirements": role_requirements,
            "required_roles": required_roles
        }

    # Return the completed project dictionary.
    return projects

def complete_role_value_text(selected_roles, input_text, default_value, value_type="float"):
    """
    Completes a role-value text input so every selected role has a value.

    Example:
    selected_roles = ["Tester", "Programmer"]
    input_text = "Tester: 4"
    default_value = 1

    returns:
    "Tester: 4, Programmer: 1"
    """

    role_values = {}

    # Parse existing text input.
    if input_text is not None and str(input_text).strip() != "":
        entries = str(input_text).split(",")

        for entry in entries:
            if ":" in entry:
                role, value = entry.split(":", 1)
                role = role.strip()

                try:
                    if value_type == "int":
                        role_values[role] = int(float(value.strip()))
                    else:
                        role_values[role] = float(value.strip())
                except ValueError:
                    pass

    # Fill missing selected roles with default.
    for role in selected_roles:
        if role not in role_values:
            role_values[role] = default_value

    # Format nicely for display/storage.
    formatted_parts = []

    for role in selected_roles:
        value = role_values[role]

        if value_type == "int":
            formatted_parts.append(f"{role}: {int(value)}")
        else:
            # Show 1 instead of 1.0, but keep decimals like 1.5
            if float(value).is_integer():
                formatted_parts.append(f"{role}: {int(value)}")
            else:
                formatted_parts.append(f"{role}: {value}")

    return ", ".join(formatted_parts)



def apply_project_defaults(project_records, default_workers_per_role=1, default_role_hours_per_week=4):
    """
    Completes project role requirement fields using the visible default values.
    This keeps imported projects and edited table rows consistent with manual entry.
    """

    completed_records = []

    for project in project_records:
        record = dict(project)
        selected_roles = split_items(record.get("Roles", ""))

        if selected_roles:
            record["Workers per role"] = complete_role_value_text(
                selected_roles=selected_roles,
                input_text=record.get("Workers per role", ""),
                default_value=default_workers_per_role,
                value_type="int"
            )

            record["Role hours per week"] = complete_role_value_text(
                selected_roles=selected_roles,
                input_text=record.get("Role hours per week", ""),
                default_value=default_role_hours_per_week,
                value_type="float"
            )

        record["Role hours/week"] = default_role_hours_per_week
        completed_records.append(record)

    return completed_records

# -----------------------------
# WORKER DATA BUILDERS
# -----------------------------

def build_workers(worker_rows):
    # List that will store worker names.
    workers = []

    # Loop through each worker row entered in the Streamlit app.
    for row in worker_rows:
        # Get the worker name from the row.
        worker_name = row["Worker name"]

        # Add the worker if the name is not blank.
        if str(worker_name).strip() != "":
            workers.append(worker_name)

    # Return the list of worker names.
    return workers

def build_worker_preferences(worker_rows):
    """
    Converts worker rows into a dictionary of preferred roles.
    Format:
    worker_preferences[worker] = [preferred_role_1, preferred_role_2, ...]
    """

    worker_preferences = {}

    for row in worker_rows:
        worker_name = str(row.get("Worker name", "")).strip()

        if worker_name != "":
            worker_preferences[worker_name] = split_items(
                row.get("Preferred roles", "")
            )
    
    return worker_preferences

# Weekly hours = estimated hours per week the worker can spend on projects.
def build_weekly_hours(worker_rows):
    """
    Converts worker rows into a dictionary of worker weekly hours.
    Weekly hours controls how many active project hours a worker can hold within the same week.
    """

    weekly_hours = {}

    for row in worker_rows:
        worker_name = row["Worker name"]

        if str(worker_name).strip() != "":
            try:
                weekly_hours[worker_name] = float(row.get("Weekly hours", 0))
            except ValueError:
                weekly_hours[worker_name] = 0
    return weekly_hours

def build_worker_unavailability(worker_rows):
    """
    Converts plain week-number entries from the Unavailability field into fully unavailable weeks.

    Date ranges are handled separately by build_worker_weekly_capacity().

    Example:
    Unavailability = "3, 4, 2026-08-10 to 2026-08-14"

    returns:
    worker_unavailability[worker] = [3, 4]
    """

    worker_unavailability = {}

    for row in worker_rows:
        worker_name = str(row.get("Worker name", "")).strip()

        if worker_name != "":
            unavailable_weeks = []

            for item in split_unavailability_items(row.get("Unavailability", "")):
                week_number = parse_week_number(item)

                if week_number is not None:
                    unavailable_weeks.append(week_number)

            worker_unavailability[worker_name] = unavailable_weeks

    return worker_unavailability

def build_worker_weekly_capacity(
    worker_rows,
    planning_start_date,
    max_weeks=52,
    workdays_per_week=5
):
    """
    Builds week-specific worker capacity after applying date unavailability.

    Plain week numbers are handled by build_worker_unavailability().
    Date ranges reduce weekly capacity based on unavailable weekdays.

    Example:
    Weekly hours = 10
    Unavailable for 2 weekdays in Week 3
    Adjusted Week 3 capacity = 6
    """

    worker_weekly_capacity = {}

    for row in worker_rows:
        worker_name = str(row.get("Worker name", "")).strip()

        if worker_name == "":
            continue

        try:
            base_weekly_hours = float(row.get("Weekly hours", 0))
        except (TypeError, ValueError):
            base_weekly_hours = 0

        # Start each week with the worker's normal weekly capacity.
        worker_weekly_capacity[worker_name] = {
            week: base_weekly_hours
            for week in range(1, max_weeks + 1)
        }

        unavailable_workdays_by_week = {}

        for item in split_unavailability_items(row.get("Unavailability", "")):
            # Plain numbers are full unavailable weeks and are handled elsewhere.
            if parse_week_number(item) is not None:
                continue

            parsed_date_range = parse_unavailability_date_item(item)

            if parsed_date_range is None:
                continue

            unavailable_start_date, unavailable_end_date = parsed_date_range

            current_date = unavailable_start_date

            while current_date <= unavailable_end_date:
                # Count Monday-Friday only.
                if current_date.weekday() < 5:
                    week_number = date_to_model_week(
                        input_date=current_date,
                        planning_start_date=planning_start_date
                    )

                    if week_number is not None and 1 <= week_number <= max_weeks:
                        if week_number not in unavailable_workdays_by_week:
                            unavailable_workdays_by_week[week_number] = set()

                        unavailable_workdays_by_week[week_number].add(current_date)

                current_date += timedelta(days=1)

        # Reduce capacity for weeks affected by date-based unavailability.
        for week_number, unavailable_dates in unavailable_workdays_by_week.items():
            unavailable_days = min(len(unavailable_dates), workdays_per_week)

            reduction_fraction = unavailable_days / workdays_per_week
            adjusted_capacity = base_weekly_hours * (1 - reduction_fraction)

            worker_weekly_capacity[worker_name][week_number] = max(
                0,
                adjusted_capacity
            )

    return worker_weekly_capacity


# -----------------------------
# Q MATRIX DATA BUILDERS
# -----------------------------

# Converts the saved Q matrix rows into the nested Q dictionary format required by the optimizer: Q[worker][project][role] = q_value.
def build_q_matrix(q_rows):
    # Empty dictionary that will become the nested Q matrix.
    Q = {}

    # Loop through each row in the saved Q matrix table.
    for row in q_rows:
        # Extract the worker, project, role, and Q value from the row.
        worker = row["Worker"]
        project = row["Project"]
        role = row["Role"]
        q_value = float(row["Q Value"])

        # If this worker is not already in Q, add them.
        if worker not in Q:
            Q[worker] = {}

        # If this project is not already under this worker, add it.
        if project not in Q[worker]:
            Q[worker][project] = {}

        # Store the Q value for this worker-project-role combination.
        Q[worker][project][role] = q_value

    # Return the completed nested Q matrix.
    return Q


# -----------------------------
# RESULTS DISPLAY HELPERS
# -----------------------------

def explain_unselected_projects(
    all_projects,
    selected_projects,
    projects_data,
    workers_data,
    q_data,
    project_dependencies,
    project_conflicts,
    project_schedule,
    min_suitability,
    weekly_hours
):
    """
    Builds simple user-facing explanations for why projects were not selected.
    These are diagnostic explanations based on project requirements, Q scores,
    deadlines, dependencies, conflicts, and scheduling.
    """

    explanations = []
    selected_set = set(selected_projects)

    def get_role_hours(project_data, role):
        """
        Gets the weekly hours required for a specific project role.
        Uses role-specific hours if available, otherwise falls back to the project's default role hours.
        """

        specific_role_hours = project_data.get("specific_role_hours", {})
        default_role_hours = project_data.get("role_hours_per_week", 1)

        try:
            return float(specific_role_hours.get(role, default_role_hours))
        except (TypeError, ValueError):
            return 1.0

    for project in all_projects:
        if project in selected_set:
            continue

        reasons = []

        project_data = projects_data.get(project, {})
        required_roles = project_data.get("required_roles", [])
        role_requirements = project_data.get("role_requirements", {})
        valid_start_weeks = project_data.get("valid_start_weeks", [])
        project_duration = project_data.get("estimated_duration_weeks", 1)

        #
        if project_data.get("deadline_feasible", True) is False:
            reasons.append(
                project_data.get(
                    "deadline_note",
                    "This project cannot finish before its deadline."
                )
            )

        # Deadline/start feasibility
        if len(valid_start_weeks) == 0:
            reasons.append(
                "No valid start week was available before the project deadline."
            )

        # Role/Q/hour feasibility
        for role in required_roles:
            workers_needed = int(role_requirements.get(role, 1))
            role_hours = get_role_hours(project_data, role)

            qualified_by_q = []
            qualified_with_enough_hours = []
            qualified_but_too_few_hours = []

            for worker in workers_data:
                q_score = (
                    q_data
                    .get(worker, {})
                    .get(project, {})
                    .get(role, 0)
                )

                if q_score >= min_suitability:
                    qualified_by_q.append(worker)

                    try:
                        worker_capacity = float(weekly_hours.get(worker, 0))
                    except (TypeError, ValueError):
                        worker_capacity = 0

                    if worker_capacity >= role_hours:
                        qualified_with_enough_hours.append(worker)
                    else:
                        qualified_but_too_few_hours.append(
                            f"{worker} ({worker_capacity:g}/{role_hours:g} hrs)"
                        )

            if len(qualified_by_q) < workers_needed:
                reasons.append(
                    f"Role '{role}' needed {workers_needed} worker(s), but "
                    f"{len(qualified_by_q)} worker(s) met the minimum Q score of "
                    f"{min_suitability}. The minimum Q score can be adjusted in the sidebar."
                )

            elif len(qualified_with_enough_hours) < workers_needed:
                examples = ", ".join(qualified_but_too_few_hours[:3])

                if len(qualified_with_enough_hours) == 0:
                    capacity_text = "but none of them had enough weekly capacity."
                else:
                    capacity_text = (
                        f"but only {len(qualified_with_enough_hours)} of them had enough weekly capacity."
                    )

                reason = (
                    f"Role '{role}' requires {role_hours:g} hour(s)/week. "
                    f"{len(qualified_by_q)} worker(s) met the minimum Q score of "
                    f"{min_suitability}, {capacity_text}"
                )

                if examples:
                    reason += f" Suitable workers below the required capacity: {examples}."

                reasons.append(reason)

        # Dependency explanations
        for dependent_project, required_project in project_dependencies:
            if dependent_project != project:
                continue

            # Case 1: required project was not selected
            if required_project not in selected_set:
                reasons.append(
                    f"It depends on '{required_project}', which was not selected."
                )
                continue

            # Case 2: required project was selected, but timing may not work
            required_schedule = project_schedule.get(required_project, {})
            required_active_weeks = required_schedule.get("Scheduled Active Weeks", [])

            if required_active_weeks:
                required_finish_week = max(required_active_weeks)
                earliest_start_after_dependency = required_finish_week + 1

                possible_start_weeks_after_dependency = [
                    week for week in valid_start_weeks
                    if week >= earliest_start_after_dependency
                ]

                if len(possible_start_weeks_after_dependency) == 0:
                    reasons.append(
                        f"This project depends on '{required_project}', but its deadline and "
                        f"{project_duration}-week duration leave no valid start week after "
                        f"'{required_project}' finishes."
                    )

        # Conflict explanations
        for project_a, project_b in project_conflicts:
            if project == project_a and project_b in selected_set:
                reasons.append(
                    f"It conflicts with selected project '{project_b}', so both projects could not fit in the selected schedule."
                )

            elif project == project_b and project_a in selected_set:
                reasons.append(
                    f"It conflicts with selected project '{project_a}', so both projects could not fit in the selected schedule."
                )

        # General fallback
        if len(reasons) == 0:
            reasons.append(
                "This project may be feasible on its own, but it did not fit with the best overall schedule when considering the other selected projects, worker capacity, deadlines, and project relationships."
            )

        explanations.append({
            "Project": project,
            "Likely reason not selected": " ".join(reasons)
        })

    return explanations

def build_weekly_hours_usage(assignments, projects_data, weekly_hours, project_schedule, worker_unavailability, worker_weekly_capacity=None):    
    """
    Builds a table showing how many weekly hours each worker uses
    compared to their available weekly hours.

    This uses the optimizer's chosen project schedule, not fixed active weeks,
    so workers are only counted during the weeks when selected projects are active.
    """

    if worker_weekly_capacity is None:
        worker_weekly_capacity = {}

    usage = {}

    # Loop through selected projects and use their scheduled active weeks.
    for project, project_assignments in assignments.items():
        active_weeks = project_schedule.get(project, {}).get("Scheduled Active Weeks", [])
        specific_role_hours = projects_data[project].get("specific_role_hours", {})
        default_role_hours = projects_data[project].get("role_hours_per_week", 1)

        for role, assigned_workers in project_assignments.items():
            role_hours = specific_role_hours.get(role, default_role_hours)

            if not isinstance(assigned_workers, list):
                assigned_workers = [assigned_workers]

            for worker in assigned_workers:
                for week in active_weeks:
                    key = (worker, week)

                    if key not in usage:
                        usage[key] = 0

                    usage[key] += role_hours

    usage_rows = []

    # Sort by week first, then worker name, so the table reads like a weekly schedule.
    for (worker, week), used_hours in sorted(
        usage.items(),
        key=lambda item: (item[0][1], item[0][0])
    ):
        if week in worker_unavailability.get(worker, []):
            available_hours = 0
        else:
            available_hours = (
                worker_weekly_capacity.get(worker, {}).get(week, weekly_hours.get(worker, 0))
            )

        usage_rows.append({
            "Week": week,
            "Worker": worker,
            "Assigned Hours": used_hours,
            "Weekly Available Hours": available_hours,
            "Remaining Weekly Hours": available_hours - used_hours
        })

    return usage_rows

# -----------------------------
# RESULTS DOWNLOAD HELPERS
# -----------------------------

def build_summary_rows(
        total_selected_projects,
        total_entered_projects,
        total_assignments,
        total_required_roles_for_selected_projects,
        objective_score
):
    """
    Build summary rows to export results.
    """

    return [
        {
            "Metric": "Selected Projects",
            "Value": f"{total_selected_projects} / {total_entered_projects}"
        },
        {
            "Metric": "Role Assignments",
            "Value": f"{total_assignments} / {total_required_roles_for_selected_projects}"
        },
        {
            "Metric": "Objective Score",
            "Value": f"{objective_score:.2f}"
        }
    ]

def build_settings_rows(
        min_suitability,
        completion_bonus_weight,
        preference_bonus_weight,
        priority_bonus_weight,
        start_week_penalty_weight
): 

    return [
        {
            "Setting": "Minimum Q score",
            "Value": min_suitability
        },
        {
            "Setting": "Project completion bonus weight",
            "Value": completion_bonus_weight
        },
        {
            "Setting": "Worker preference bonus weight",
            "Value": preference_bonus_weight
        },
        {
            "Setting": "Project priority bonus weight",
            "Value": priority_bonus_weight
        },
        {
            "Setting": "Start week penalty weight",
            "Value": start_week_penalty_weight
        }
    ]

def create_results(
        summary_rows,
        settings_rows,
        schedule_rows,
        assignment_rows,
        unselected_rows,
        workload_rows,
        deadline_rows,
        relationship_rows
):
    """
    Creates a downloadable Excel file containing the results tables.
    Applies basic formatting so the workbook is easier to read.
    """

    from io import BytesIO
    import re
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    output = BytesIO()

    sheets = {
        "Summary": summary_rows,
        "Settings": settings_rows,
        "Schedule": schedule_rows,
        "Assignments": assignment_rows,
        "Unselected Projects": unselected_rows,
        "Worker Workload": workload_rows,
        "Deadlines": deadline_rows,
        "Relationships": relationship_rows
    }

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Write each table to its own sheet.
        for sheet_name, rows in sheets.items():
            pd.DataFrame(rows).to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3")
        )

        for sheet_name in sheets.keys():
            worksheet = workbook[sheet_name]

            max_row = worksheet.max_row
            max_col = worksheet.max_column

            if max_row == 1 and max_col == 1 and worksheet["A1"].value is None:
                continue

            # Freeze the header row.
            worksheet.freeze_panes = "A2"

            # Style the header row.
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            # Style body cells.
            for row in worksheet.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Add an Excel table if the sheet has data.
            if max_row >= 2 and max_col >= 1:
                table_ref = f"A1:{get_column_letter(max_col)}{max_row}"

                safe_name = re.sub(r"[^A-Za-z0-9_]", "", sheet_name)
                table_name = f"{safe_name}Table"

                table = Table(displayName=table_name, ref=table_ref)

                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )

                table.tableStyleInfo = style
                worksheet.add_table(table)

            # Set sensible column widths.
            for col_num in range(1, max_col + 1):
                column_letter = get_column_letter(col_num)

                max_length = 0
                for cell in worksheet[column_letter]:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))

                adjusted_width = min(max(max_length + 2, 12), 45)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Make long-text sheets easier to read.
            if sheet_name == "Unselected Projects":
                worksheet.column_dimensions["A"].width = 28
                worksheet.column_dimensions["B"].width = 110

            if sheet_name == "Schedule":
                worksheet.column_dimensions["A"].width = 32
                worksheet.column_dimensions["E"].width = 30
                worksheet.column_dimensions["F"].width = 42

            if sheet_name == "Assignments":
                worksheet.column_dimensions["A"].width = 32
                worksheet.column_dimensions["B"].width = 22
                worksheet.column_dimensions["C"].width = 28

            if sheet_name == "Worker Workload":
                worksheet.column_dimensions["A"].width = 12
                worksheet.column_dimensions["B"].width = 20

            # Center numeric-looking columns.
            for row in worksheet.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="center", vertical="top")

            # Row heights for readability.
            worksheet.row_dimensions[1].height = 24

            if sheet_name == "Unselected Projects":
                for row_num in range(2, max_row + 1):
                    worksheet.row_dimensions[row_num].height = 65

            elif sheet_name in ["Schedule", "Assignments", "Relationships"]:
                for row_num in range(2, max_row + 1):
                    worksheet.row_dimensions[row_num].height = 28

            else:
                for row_num in range(2, max_row + 1):
                    worksheet.row_dimensions[row_num].height = 22

        # Make the Summary sheet a little cleaner.
        if "Summary" in workbook.sheetnames:
            summary_sheet = workbook["Summary"]
            summary_sheet.column_dimensions["A"].width = 28
            summary_sheet.column_dimensions["B"].width = 18

            for cell in summary_sheet["B"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    output.seek(0)
    return output.getvalue()